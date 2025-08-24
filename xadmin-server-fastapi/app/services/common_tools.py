"""
通用工具服务
"""
import io
import qrcode
import base64
import platform
import psutil
import sys
from typing import Dict, Any, Optional
from PIL import Image
from sqlalchemy.orm import Session
from app.core.config import settings
from app.core.database import engine
import redis
import logging

logger = logging.getLogger(__name__)


class SystemInfoService:
    """系统信息服务"""
    
    @staticmethod
    def get_system_info() -> Dict[str, Any]:
        """获取系统信息"""
        try:
            # 系统信息
            system_info = {
                'platform': platform.system(),
                'platform_version': platform.version(),
                'architecture': platform.machine(),
                'hostname': platform.node(),
                'processor': platform.processor()
            }
            
            # Python信息
            python_info = {
                'version': sys.version,
                'implementation': platform.python_implementation(),
                'compiler': platform.python_compiler()
            }
            
            # 内存信息
            memory = psutil.virtual_memory()
            memory_info = {
                'total': f"{memory.total // (1024**3)} GB",
                'available': f"{memory.available // (1024**3)} GB",
                'percent': memory.percent,
                'used': f"{memory.used // (1024**3)} GB"
            }
            
            # 磁盘信息
            disk = psutil.disk_usage('/')
            disk_info = {
                'total': f"{disk.total // (1024**3)} GB",
                'used': f"{disk.used // (1024**3)} GB",
                'free': f"{disk.free // (1024**3)} GB",
                'percent': (disk.used / disk.total) * 100
            }
            
            # CPU信息
            cpu_info = {
                'physical_cores': psutil.cpu_count(logical=False),
                'total_cores': psutil.cpu_count(logical=True),
                'max_frequency': f"{psutil.cpu_freq().max:.2f} MHz" if psutil.cpu_freq() else "N/A",
                'current_frequency': f"{psutil.cpu_freq().current:.2f} MHz" if psutil.cpu_freq() else "N/A",
                'cpu_usage': f"{psutil.cpu_percent(interval=1)}%"
            }
            
            return {
                'system': system_info,
                'python': python_info,
                'memory': memory_info,
                'disk': disk_info,
                'cpu': cpu_info
            }
            
        except Exception as e:
            logger.error(f"Failed to get system info: {e}")
            return {
                'error': str(e)
            }
    
    @staticmethod
    def check_health() -> Dict[str, str]:
        """健康检查"""
        health_status = {
            'status': 'healthy',
            'database': 'unknown',
            'redis': 'unknown'
        }
        
        # 检查数据库连接
        try:
            with engine.connect() as conn:
                conn.execute("SELECT 1")
            health_status['database'] = 'connected'
        except Exception as e:
            health_status['database'] = f'error: {str(e)}'
            health_status['status'] = 'unhealthy'
        
        # 检查Redis连接
        try:
            redis_client = redis.from_url(settings.REDIS_URL)
            redis_client.ping()
            health_status['redis'] = 'connected'
        except Exception as e:
            health_status['redis'] = f'error: {str(e)}'
            # Redis不是必需的，所以不影响整体健康状态
        
        return health_status


class QRCodeService:
    """二维码服务"""
    
    @staticmethod
    def generate_qrcode(
        content: str,
        size: int = 200,
        border: int = 4,
        error_correction: str = 'M'
    ) -> Dict[str, Any]:
        """生成二维码"""
        try:
            # 错误纠正级别映射
            error_correction_map = {
                'L': qrcode.constants.ERROR_CORRECT_L,
                'M': qrcode.constants.ERROR_CORRECT_M,
                'Q': qrcode.constants.ERROR_CORRECT_Q,
                'H': qrcode.constants.ERROR_CORRECT_H
            }
            
            # 创建二维码实例
            qr = qrcode.QRCode(
                version=1,
                error_correction=error_correction_map.get(error_correction, qrcode.constants.ERROR_CORRECT_M),
                box_size=10,
                border=border
            )
            
            # 添加数据
            qr.add_data(content)
            qr.make(fit=True)
            
            # 创建图片
            img = qr.make_image(fill_color="black", back_color="white")
            
            # 调整图片大小
            img = img.resize((size, size), Image.Resampling.LANCZOS)
            
            # 转换为base64
            buffer = io.BytesIO()
            img.save(buffer, format='PNG')
            img_base64 = base64.b64encode(buffer.getvalue()).decode()
            
            return {
                'qrcode_image': f"data:image/png;base64,{img_base64}",
                'content': content,
                'size': size,
                'border': border,
                'error_correction': error_correction
            }
            
        except Exception as e:
            logger.error(f"Failed to generate QR code: {e}")
            raise ValueError(f"生成二维码失败: {str(e)}")


class DataExportService:
    """数据导出服务"""
    
    def __init__(self, db: Session):
        self.db = db
    
    def export_data(
        self,
        table_name: str,
        export_type: str = 'xlsx',
        fields: Optional[list] = None,
        filters: Optional[dict] = None,
        limit: Optional[int] = None
    ) -> bytes:
        """导出数据"""
        try:
            # 构建查询
            query = self._build_export_query(table_name, fields, filters, limit)
            
            # 执行查询
            results = self.db.execute(query).fetchall()
            
            # 根据导出类型生成文件
            if export_type == 'xlsx':
                return self._export_to_xlsx(results, fields)
            elif export_type == 'csv':
                return self._export_to_csv(results, fields)
            elif export_type == 'pdf':
                return self._export_to_pdf(results, fields)
            else:
                raise ValueError(f"不支持的导出类型: {export_type}")
                
        except Exception as e:
            logger.error(f"Failed to export data: {e}")
            raise ValueError(f"数据导出失败: {str(e)}")
    
    def _build_export_query(self, table_name: str, fields: list, filters: dict, limit: int) -> str:
        """构建导出查询"""
        # 这里应该根据实际的表结构和需求来构建查询
        # 为了安全起见，应该验证table_name和fields
        
        # 基础查询
        if fields:
            fields_str = ', '.join(fields)
        else:
            fields_str = '*'
        
        query = f"SELECT {fields_str} FROM {table_name}"
        
        # 添加过滤条件
        if filters:
            conditions = []
            for key, value in filters.items():
                if isinstance(value, str):
                    conditions.append(f"{key} LIKE '%{value}%'")
                else:
                    conditions.append(f"{key} = {value}")
            
            if conditions:
                query += " WHERE " + " AND ".join(conditions)
        
        # 添加限制
        if limit:
            query += f" LIMIT {limit}"
        
        return query
    
    def _export_to_xlsx(self, results: list, fields: list) -> bytes:
        """导出为Excel格式"""
        try:
            import openpyxl
            from io import BytesIO
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "导出数据"
            
            # 写入表头
            if fields:
                for col, field in enumerate(fields, 1):
                    ws.cell(row=1, column=col, value=field)
            
            # 写入数据
            for row_idx, row_data in enumerate(results, 2):
                for col_idx, value in enumerate(row_data, 1):
                    ws.cell(row=row_idx, column=col_idx, value=value)
            
            # 保存到内存
            buffer = BytesIO()
            wb.save(buffer)
            return buffer.getvalue()
            
        except ImportError:
            raise ValueError("openpyxl库未安装，无法导出Excel格式")
        except Exception as e:
            raise ValueError(f"Excel导出失败: {str(e)}")
    
    def _export_to_csv(self, results: list, fields: list) -> bytes:
        """导出为CSV格式"""
        try:
            import csv
            from io import StringIO, BytesIO
            
            output = StringIO()
            writer = csv.writer(output)
            
            # 写入表头
            if fields:
                writer.writerow(fields)
            
            # 写入数据
            writer.writerows(results)
            
            # 转换为bytes
            return output.getvalue().encode('utf-8-sig')  # 添加BOM以支持中文
            
        except Exception as e:
            raise ValueError(f"CSV导出失败: {str(e)}")
    
    def _export_to_pdf(self, results: list, fields: list) -> bytes:
        """导出为PDF格式"""
        # 这里可以使用reportlab等库来生成PDF
        # 由于复杂性，这里只是一个示例
        raise NotImplementedError("PDF导出功能暂未实现")


class DataImportService:
    """数据导入服务"""
    
    def __init__(self, db: Session):
        self.db = db
    
    def preview_import(self, file_path: str, table_name: str) -> Dict[str, Any]:
        """预览导入数据"""
        try:
            # 读取文件
            data = self._read_import_file(file_path)
            
            # 验证数据
            validation_result = self._validate_import_data(data, table_name)
            
            return {
                'action': 'preview',
                'total_rows': len(data),
                'valid_rows': validation_result['valid_count'],
                'invalid_rows': validation_result['invalid_count'],
                'columns': validation_result['columns'],
                'sample_data': data[:5],  # 显示前5行作为示例
                'errors': validation_result['errors']
            }
            
        except Exception as e:
            logger.error(f"Failed to preview import: {e}")
            raise ValueError(f"预览导入失败: {str(e)}")
    
    def import_data(self, file_path: str, table_name: str, mapping: dict = None) -> Dict[str, Any]:
        """导入数据"""
        try:
            # 读取文件
            data = self._read_import_file(file_path)
            
            # 应用字段映射
            if mapping:
                data = self._apply_field_mapping(data, mapping)
            
            # 执行导入
            result = self._execute_import(data, table_name)
            
            return {
                'action': 'import',
                'total_rows': len(data),
                'success_rows': result['success_count'],
                'failed_rows': result['failed_count'],
                'errors': result['errors']
            }
            
        except Exception as e:
            logger.error(f"Failed to import data: {e}")
            raise ValueError(f"数据导入失败: {str(e)}")
    
    def _read_import_file(self, file_path: str) -> list:
        """读取导入文件"""
        if file_path.endswith('.xlsx') or file_path.endswith('.xls'):
            return self._read_excel_file(file_path)
        elif file_path.endswith('.csv'):
            return self._read_csv_file(file_path)
        else:
            raise ValueError("不支持的文件格式")
    
    def _read_excel_file(self, file_path: str) -> list:
        """读取Excel文件"""
        try:
            import openpyxl
            
            wb = openpyxl.load_workbook(file_path)
            ws = wb.active
            
            data = []
            for row in ws.iter_rows(values_only=True):
                data.append(list(row))
            
            return data
            
        except ImportError:
            raise ValueError("openpyxl库未安装，无法读取Excel文件")
        except Exception as e:
            raise ValueError(f"读取Excel文件失败: {str(e)}")
    
    def _read_csv_file(self, file_path: str) -> list:
        """读取CSV文件"""
        try:
            import csv
            
            data = []
            with open(file_path, 'r', encoding='utf-8-sig') as f:
                reader = csv.reader(f)
                for row in reader:
                    data.append(row)
            
            return data
            
        except Exception as e:
            raise ValueError(f"读取CSV文件失败: {str(e)}")
    
    def _validate_import_data(self, data: list, table_name: str) -> Dict[str, Any]:
        """验证导入数据"""
        # 这里应该根据实际的表结构来验证数据
        # 简化示例
        
        valid_count = 0
        invalid_count = 0
        errors = []
        columns = data[0] if data else []
        
        for idx, row in enumerate(data[1:], 2):  # 跳过表头
            try:
                # 这里添加具体的验证逻辑
                valid_count += 1
            except Exception as e:
                invalid_count += 1
                errors.append({
                    'row': idx,
                    'error': str(e)
                })
        
        return {
            'valid_count': valid_count,
            'invalid_count': invalid_count,
            'columns': columns,
            'errors': errors
        }
    
    def _apply_field_mapping(self, data: list, mapping: dict) -> list:
        """应用字段映射"""
        if not data:
            return data
        
        headers = data[0]
        mapped_data = [data[0]]  # 保留原始表头
        
        # 应用映射
        for row in data[1:]:
            mapped_row = []
            for i, value in enumerate(row):
                if i < len(headers):
                    field_name = headers[i]
                    mapped_field = mapping.get(field_name, field_name)
                    mapped_row.append(value)
            mapped_data.append(mapped_row)
        
        return mapped_data
    
    def _execute_import(self, data: list, table_name: str) -> Dict[str, Any]:
        """执行数据导入"""
        # 这里应该根据实际需求实现数据导入逻辑
        # 简化示例
        
        success_count = 0
        failed_count = 0
        errors = []
        
        try:
            # 这里添加具体的导入逻辑
            success_count = len(data) - 1  # 减去表头
            
        except Exception as e:
            failed_count = len(data) - 1
            errors.append({
                'error': str(e)
            })
        
        return {
            'success_count': success_count,
            'failed_count': failed_count,
            'errors': errors
        }